"""Beer-pour-machine RFID verification test logger.

Drives the existing ``rfid_gc_live`` C binary one trial at a time and records
every trial to a single master Excel workbook (``beer-pour-results.xlsx``).

The C binary is the arbitrated dual-antenna scanner from ``rfid_gc_live.c``:
it runs both antennas at maximum scan rate, accumulates per-EPC stats for one
decision window, and at every window boundary emits one line on stdout:

- ``[]`` when no tag was decisively attributed in that window
- ``[TX=N mW] [(0)(-XX.X) EPC, (1)(-YY.Y) EPC]`` otherwise (ANSI-coloured),
  with antenna 0 in the left slot, antenna 1 in the right slot. Empty slots
  render as pure whitespace so the comma column never shifts.

The harness does NOT ask the operator which antenna the cup is going to be
slid over -- in real-world use the system has no advance knowledge of that.
Instead it lets the arbitrator make the call and then characterises the
quality of that call. The results spreadsheet is intentionally compact;
each trial is one row with these columns:

- ``session_id``           timestamp of the test session
- ``scenario``             physical setup under test
- ``power_mw``             TX power used
- ``scans_in_3s``          total CAENRFID_InventoryTag attempts both
                           antennas completed in the first 3 s of the
                           trial (the work the reader did to produce
                           the result on this row)
- ``winning_antenna``      per-EPC home antenna, e.g. ``6E6F76 -> ant0``
- ``cross_reads``          number of EPCs seen on BOTH antennas (same-tag
                           leakage; 0 is the happy path)
- ``best_rssi_winner_dbm`` strongest RSSI on the winning antenna
- ``detected_epcs``        every distinct EPC seen during the trial
- ``tag_photo``            embedded thumbnail of the tag used
- ``result``               PASS / SLOW / DIRTY / FAIL one-word verdict
- ``notes``                operator comment entered after the trial

After every trial the operator is asked whether to keep that row -- press
ENTER (or 'y') to append, 'n' to discard -- and then optionally enter a
free-text comment that goes into ``notes``.

Usage::

    python3 beer_pour_logger.py
"""

from __future__ import annotations

import datetime as dt
import os
import re
import signal
import subprocess
import sys
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


# ────────────────────────── paths & config ──────────────────────────


SCRIPT_DIR     = Path(__file__).resolve().parent
RFID_BINARY    = SCRIPT_DIR / "rfid_gc_live"
IMAGES_DIR     = SCRIPT_DIR / "images"
TAGS_DIR       = IMAGES_DIR / "tags"
SCENARIOS_DIR  = IMAGES_DIR / "scenarios"
RESULTS_XLSX   = SCRIPT_DIR / "beer-pour-results.xlsx"
THUMB_DIR      = SCRIPT_DIR / ".thumbs"

# Fixed set of physical scenarios under test on the beer-pour machine.
# Extend this list (and drop a matching <name>.png into images/scenarios/)
# when you add a scenario; everything else picks it up automatically.
SCENARIOS: List[str] = [
    "drip_tray_empty_cup_empty",
    "drip_tray_half_full_cup_empty",
]

# Power levels (mW) we sweep through. Passed verbatim as the positional
# argument to rfid_gc_live.
POWER_LEVELS_MW: List[int] = [30, 175, 316]

# The reader's decision window is 1000 ms, so each trial captures
# roughly TRIAL_DURATION_S window lines. 5 s gives 5 windows -- enough
# to comfortably detect a tag that needs to verify in 3 s and still
# observe a couple of windows of behaviour after that.
TRIAL_DURATION_S: float = 5.0

# A tag must be attributed within this many seconds for the trial to
# pass. Matches the product spec: "verification happens within 3 s of
# the user sliding a tagged cup over the antenna".
VERIFY_DEADLINE_S: float = 3.0

# Photo thumbnail height (pixels) embedded in the workbook. Row height
# is in points; one row comfortably holds an 80-px-tall thumbnail.
THUMB_HEIGHT_PX  = 80
ROW_HEIGHT_PT    = 64


# ──────────────────────── output line parsing ───────────────────────
#
# rfid_gc_live emits its sweep lines with ANSI colour codes (cyan for the
# [S0=N S1=N] / [TX=...] prefixes, yellow for the (N) antenna labels,
# green/red for the EPCs). We strip the colours first, then run a small
# set of regexes on the plain text. The line shapes after stripping are:
#
#   "[GC] Ready. ..."                                 -> start of trial
#   "[S0=137 S1=138] []"                              -> empty window
#   "[S0=137 S1=138] [TX=30 mW] [(0)(-58.3) EPC0,   (1)(-61.7) EPC1]"
#   "[S0=137 S1=138] [TX=30 mW] [(0)(-58.3) EPC0,                  ]"
#   "[S0=137 S1=138] [TX=30 mW] [                ,   (1)(-61.7) EPC1]"
#
# Slot padding is just trailing whitespace, so strip() on each slot string
# is enough; the comma is the only separator.
#
# The S0/S1 scan counters report how many CAENRFID_InventoryTag calls
# the C binary completed on each antenna during the 1-second decision
# window that produced the line. They reset at every window boundary.

ANSI_RE          = re.compile(r"\x1b\[[0-9;]*m")
READY_RE         = re.compile(r"\[GC\]\s+Ready\.")
SCANS_PREFIX_RE  = re.compile(r"^\s*\[S0=(?P<s0>\d+)\s+S1=(?P<s1>\d+)\]\s*")
EMPTY_WINDOW_RE  = re.compile(r"^\s*\[\]\s*$")
WINDOW_RE        = re.compile(
    r"^\s*\[TX=(?P<power>\d+)\s*mW\]\s+\[(?P<inner>.*)\]\s*$"
)
SLOT_HEADER_RE   = re.compile(r"^\((?P<ant>\d+)\)(?P<tags>.+)$")
TAG_PAIR_RE      = re.compile(r"\((?P<rssi>-?\d+\.\d+)\)\s+(?P<epc>[0-9A-Fa-f]+)")


def _parse_slot(slot: str) -> List[Tuple[str, float]]:
    """Return [(EPC, rssi_dbm), ...] for one slot, or [] if it's empty."""
    if not slot:
        return []
    m = SLOT_HEADER_RE.match(slot)
    if not m:
        return []
    return [
        (epc, float(rssi))
        for rssi, epc in TAG_PAIR_RE.findall(m.group("tags"))
    ]


@dataclass
class WindowRead:
    """One decision-window line from the C binary, with arrival timestamp.

    ``t_offset_s`` is seconds since the trial's ``[GC] Ready`` line.
    ``scans_ant0`` / ``scans_ant1`` come from the ``[S0=N S1=N]`` prefix
    the C binary emits on every window line -- they're the number of
    CAENRFID_InventoryTag calls completed on each antenna during this
    decision window."""

    idx: int
    t_offset_s: float
    raw: str
    is_empty: bool
    scans_ant0: int = 0
    scans_ant1: int = 0
    ant0: List[Tuple[str, float]] = field(default_factory=list)
    ant1: List[Tuple[str, float]] = field(default_factory=list)

    def tags_on(self, antenna: int) -> List[Tuple[str, float]]:
        return self.ant0 if antenna == 0 else self.ant1


def parse_window_line(idx: int, t_offset_s: float, line: str) -> Optional[WindowRead]:
    """Turn one stdout line from rfid_gc_live into a WindowRead, or None
    if the line isn't a sweep line at all (e.g. the Ready banner)."""
    clean = ANSI_RE.sub("", line).rstrip("\r\n")

    # Pull the [S0=N S1=N] prefix off the front, if present. We tolerate
    # its absence so trials run against an older C binary still produce
    # parseable rows (scans_ant0/scans_ant1 will just stay at 0).
    scans_ant0 = 0
    scans_ant1 = 0
    pm = SCANS_PREFIX_RE.match(clean)
    if pm:
        scans_ant0 = int(pm.group("s0"))
        scans_ant1 = int(pm.group("s1"))
        clean = clean[pm.end():]

    if EMPTY_WINDOW_RE.match(clean):
        return WindowRead(idx=idx, t_offset_s=t_offset_s, raw=clean,
                          is_empty=True,
                          scans_ant0=scans_ant0, scans_ant1=scans_ant1)
    m = WINDOW_RE.match(clean)
    if not m:
        return None
    inner = m.group("inner")
    if "," not in inner:
        return None
    left, right = inner.split(",", 1)
    return WindowRead(
        idx=idx,
        t_offset_s=t_offset_s,
        raw=clean,
        is_empty=False,
        scans_ant0=scans_ant0,
        scans_ant1=scans_ant1,
        ant0=_parse_slot(left.strip()),
        ant1=_parse_slot(right.strip()),
    )


# ─────────────────────────── data types ────────────────────────────


@dataclass
class TrialResult:
    scenario: str
    power_mw: int
    tag: str
    trial_num: int
    start_time: dt.datetime          # wall-clock at [GC] Ready
    duration_s: float                # actual measured trial length
    windows: List[WindowRead] = field(default_factory=list)

    # ----- derived metrics ---------------------------------------------------

    @property
    def n_windows(self) -> int:
        return len(self.windows)

    @property
    def detected_epcs(self) -> List[str]:
        seen: List[str] = []
        for w in self.windows:
            for epc, _ in w.ant0 + w.ant1:
                if epc not in seen:
                    seen.append(epc)
        return seen

    def _epc_antenna_hit_counts(self) -> Dict[str, Tuple[int, int]]:
        """Per EPC: (windows attributed on ant0, windows attributed on ant1)."""
        counts: Dict[str, List[int]] = {}
        for w in self.windows:
            for ant, tags in ((0, w.ant0), (1, w.ant1)):
                for epc, _ in tags:
                    if epc not in counts:
                        counts[epc] = [0, 0]
                    counts[epc][ant] += 1
        return {epc: (c[0], c[1]) for epc, c in counts.items()}

    @property
    def epc_home_antennas(self) -> Dict[str, int]:
        """For each detected EPC, the antenna it was attributed to most
        often during the trial (= where the arbitrator thinks that tag
        lives). Ties go to the lower antenna index."""
        homes: Dict[str, int] = {}
        for epc, (n0, n1) in self._epc_antenna_hit_counts().items():
            if n0 == 0 and n1 == 0:
                continue
            homes[epc] = 0 if n0 >= n1 else 1
        return homes

    @property
    def winning_antenna(self) -> Optional[int]:
        """Summary antenna for the trial. With a single detected EPC this
        is that tag's home antenna. With multiple EPCs it is the antenna
        that saw the most attributed windows overall. ``None`` if the
        trial saw no attribution."""
        if not self.detected_epcs:
            return None
        if len(self.detected_epcs) == 1:
            return self.epc_home_antennas.get(self.detected_epcs[0])
        n0 = sum(1 for w in self.windows if w.ant0)
        n1 = sum(1 for w in self.windows if w.ant1)
        if n0 == 0 and n1 == 0:
            return None
        return 0 if n0 >= n1 else 1

    @property
    def winning_antenna_per_tag(self) -> str:
        """Human-readable summary of which antenna each detected EPC was
        attributed to most often during the trial. With two cups present
        you get e.g. ``"6E6F76 -> ant0, 6E6FD6 -> ant1"`` (only the last
        6 hex chars of the EPC are shown to keep the cell compact). With
        a single EPC you get ``"<last6> -> antN"``. Returns an empty
        string when no EPC was attributed."""
        homes = self.epc_home_antennas
        if not homes:
            return ""
        return ", ".join(
            f"{epc[-6:]} -> ant{ant}"
            for epc, ant in sorted(homes.items())
        )

    def _scans_in(self, antenna: int, deadline_s: float) -> int:
        """Sum the per-window scan counts on ``antenna`` across every
        window that closed within ``deadline_s`` seconds of trial start.
        Returns 0 if the trial captured no windows within the window."""
        total = 0
        for w in self.windows:
            if w.t_offset_s > deadline_s:
                continue
            total += w.scans_ant0 if antenna == 0 else w.scans_ant1
        return total

    @property
    def scans_ant0_in_3s(self) -> int:
        """Number of CAENRFID_InventoryTag calls completed on antenna 0
        during the first VERIFY_DEADLINE_S seconds of the trial."""
        return self._scans_in(0, VERIFY_DEADLINE_S)

    @property
    def scans_ant1_in_3s(self) -> int:
        """Number of CAENRFID_InventoryTag calls completed on antenna 1
        during the first VERIFY_DEADLINE_S seconds of the trial."""
        return self._scans_in(1, VERIFY_DEADLINE_S)

    @property
    def scans_in_3s(self) -> int:
        """Total scan attempts both antennas made inside the 3-second
        verification window -- a direct measure of how much work the
        reader did before arriving at the result on this row."""
        return self.scans_ant0_in_3s + self.scans_ant1_in_3s

    @property
    def ttv_s(self) -> Optional[float]:
        """Time (since trial start) of the first window that put any tag
        on any antenna. ``None`` if it never happened."""
        for w in self.windows:
            if w.ant0 or w.ant1:
                return round(w.t_offset_s, 3)
        return None

    @property
    def verified(self) -> bool:
        return self.ttv_s is not None

    @property
    def within_deadline(self) -> bool:
        return self.ttv_s is not None and self.ttv_s <= VERIFY_DEADLINE_S

    @property
    def n_hits_winner(self) -> int:
        if self.winning_antenna is None:
            return 0
        return sum(1 for w in self.windows if w.tags_on(self.winning_antenna))

    @property
    def cross_read_epcs(self) -> List[str]:
        """EPCs that were attributed to BOTH antennas at some point during
        the trial. This is the real cross-read failure mode: the same tag
        leaking onto the wrong antenna. Two *different* tags sitting on
        two different antennas (one EPC per antenna only) is normal and
        does not appear here."""
        leaked: List[str] = []
        for epc, (n0, n1) in self._epc_antenna_hit_counts().items():
            if n0 > 0 and n1 > 0:
                leaked.append(epc)
        return sorted(leaked)

    @property
    def cross_reads(self) -> int:
        """How many distinct EPCs leaked onto more than one antenna."""
        return len(self.cross_read_epcs)

    @property
    def clean(self) -> bool:
        return self.cross_reads == 0

    @property
    def best_rssi_winner(self) -> Optional[float]:
        """Strongest (closest-to-zero) RSSI we ever saw on the winning
        antenna across the trial, or None if nothing was attributed."""
        if self.winning_antenna is None:
            return None
        best: Optional[float] = None
        for w in self.windows:
            for _, rssi in w.tags_on(self.winning_antenna):
                if best is None or rssi > best:
                    best = rssi
        return best

    @property
    def result(self) -> str:
        """One-word verdict for the summary column:

        - ``PASS``  verified, within 3 s, no cross-reads
        - ``SLOW``  verified + clean but took longer than 3 s
        - ``DIRTY`` verified but at least one EPC was attributed to more
                    than one antenna during the trial (same-tag leakage)
        - ``FAIL``  never got any attribution
        """
        if not self.verified:
            return "FAIL"
        if not self.clean:
            return "DIRTY"
        if not self.within_deadline:
            return "SLOW"
        return "PASS"


# ─────────────────────────── workbook I/O ───────────────────────────


# The trial sheet exposes a deliberately small, operator-focused column
# set. Order here is the order they appear in the spreadsheet.
TRIALS_HEADERS: List[str] = [
    "session_id",
    "scenario",
    "power_mw",
    "scans_in_3s",
    "winning_antenna",
    "cross_reads",
    "best_rssi_winner_dbm",
    "detected_epcs",
    "tag_photo",
    "result",
    "notes",
]

# Column widths in Excel character units, paired by index with the
# headers list above.
TRIALS_WIDTHS: List[int] = [
    18,   # session_id
    34,   # scenario
    11,   # power_mw
    14,   # scans_in_3s
    28,   # winning_antenna
    13,   # cross_reads
    20,   # best_rssi_winner_dbm
    54,   # detected_epcs
    22,   # tag_photo
    12,   # result
    36,   # notes
]

# Columns rendered with a centred alignment (counts / scalars). Everything
# else is left-aligned with wrapping.
_CENTERED_TRIALS = {
    "power_mw", "scans_in_3s", "cross_reads",
    "best_rssi_winner_dbm", "result",
}


# ────────────────────────── visual palette ──────────────────────────
#
# A sober, modern palette inspired by editorial dashboards: dark navy
# header, off-white striping for legibility, pill-style verdict cells.
# All colours are kept very close to neutral so the spreadsheet doesn't
# scream when projected on a meeting room screen.

FONT_FAMILY     = "Calibri"
TEXT_COLOR      = "1F2937"   # near-black with a hint of warmth
MUTED_TEXT      = "475569"
HEADER_BG       = "0F172A"   # very dark slate / navy
HEADER_FG       = "FFFFFF"
STRIPE_EVEN_BG  = "FFFFFF"
STRIPE_ODD_BG   = "F8FAFC"
GRID_COLOR      = "E5E7EB"

# Soft pastel pill colours for the verdict column. Light fill + a
# saturated dark text colour keeps them legible at small sizes.
RESULT_FILLS = {
    "PASS":  PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "SLOW":  PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "DIRTY": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
    "FAIL":  PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}
RESULT_FONTS = {
    "PASS":  Font(name=FONT_FAMILY, bold=True, color="065F46", size=11),
    "SLOW":  Font(name=FONT_FAMILY, bold=True, color="92400E", size=11),
    "DIRTY": Font(name=FONT_FAMILY, bold=True, color="9A3412", size=11),
    "FAIL":  Font(name=FONT_FAMILY, bold=True, color="991B1B", size=11),
}


def _apply_header_style(ws: Any) -> None:
    """Style the header row: bold white-on-navy, centred, wrapped, with
    a hairline bottom border. Also freezes the header row so it stays
    visible when scrolling, and disables Excel's gridlines so the
    custom striping reads cleanly."""
    header_font = Font(name=FONT_FAMILY, bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG,
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    bottom_border = Border(bottom=Side(style="thin", color=HEADER_BG))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = bottom_border
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    # Hide native gridlines so our striped rows + soft separators look
    # intentional. openpyxl's API on newer versions is sheet_view.showGridLines.
    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass


def _style_data_row(ws: Any, row: int, headers: List[str],
                    centered_cols: set, result_col: Optional[int] = None,
                    result_value: Optional[str] = None) -> None:
    """Apply the full body-row look: alternating stripe fill, hairline
    horizontal separators only (no busy vertical lines), Calibri 10 in
    near-black, centred or wrapped-left alignment depending on the
    column, and a pill-style fill+colour on the ``result`` cell if any."""
    stripe_fill = PatternFill(
        start_color=STRIPE_ODD_BG if (row % 2 == 0) else STRIPE_EVEN_BG,
        end_color  =STRIPE_ODD_BG if (row % 2 == 0) else STRIPE_EVEN_BG,
        fill_type="solid",
    )
    body_font = Font(name=FONT_FAMILY, size=10, color=TEXT_COLOR)
    body_border = Border(bottom=Side(style="hair", color=GRID_COLOR))
    centered = Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center",
                         wrap_text=True, indent=1)

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = body_font
        cell.fill = stripe_fill
        cell.border = body_border
        cell.alignment = centered if name in centered_cols else left

    if result_col is not None and result_value in RESULT_FILLS:
        cell = ws.cell(row=row, column=result_col)
        cell.fill = RESULT_FILLS[result_value]
        cell.font = RESULT_FONTS[result_value]
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=False)


def _sheet_headers(ws: Any) -> List[str]:
    """Read the header row from a sheet and return the column names in
    the order they actually appear in the file. Trailing ``None``s are
    trimmed (openpyxl pads to ``max_column``)."""
    headers = [
        ws.cell(row=1, column=c + 1).value
        for c in range(ws.max_column)
    ]
    while headers and (headers[-1] is None or headers[-1] == ""):
        headers.pop()
    return [str(h) for h in headers]


def _create_fresh_workbook() -> None:
    """Write a brand-new ``beer-pour-results.xlsx`` with the current schema and
    the styled header row applied."""
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    trials = wb.create_sheet("Trials")
    trials.append(TRIALS_HEADERS)
    for col, w in enumerate(TRIALS_WIDTHS, start=1):
        trials.column_dimensions[get_column_letter(col)].width = w
    _apply_header_style(trials)
    wb.save(RESULTS_XLSX)


def ensure_workbook() -> None:
    """Make sure ``beer-pour-results.xlsx`` exists with the current column schema.

    Schema policy:
    - File missing -> create a fresh one.
    - File present and headers match the current schema exactly ->
      re-apply header styling (in case it was stripped) and return.
    - File present but headers don't match -> archive the old file to
      ``results_archive_<timestamp>.xlsx`` and start a fresh file with
      the current schema. Old data is preserved alongside, just not
      mingled with the new layout.
    """
    if not RESULTS_XLSX.exists():
        _create_fresh_workbook()
        return

    try:
        wb = load_workbook(RESULTS_XLSX)
    except Exception as exc:
        print(f"NOTE: could not open existing {RESULTS_XLSX.name} ({exc}); "
              f"creating a fresh workbook.")
        _create_fresh_workbook()
        return

    headers_ok = (
        "Trials" in wb.sheetnames
        and _sheet_headers(wb["Trials"]) == TRIALS_HEADERS
    )

    if headers_ok:
        _apply_header_style(wb["Trials"])
        wb.save(RESULTS_XLSX)
        return

    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    archive = RESULTS_XLSX.with_name(f"beer-pour-results_archive_{stamp}.xlsx")
    print(
        f"NOTE: {RESULTS_XLSX.name} has an older column layout; "
        f"archiving it to {archive.name} and starting a fresh sheet."
    )
    RESULTS_XLSX.rename(archive)
    _create_fresh_workbook()


def _thumb_for(src_dir: Path, name: str) -> Optional[Path]:
    """Return a cached thumbnail of ``<src_dir>/<name>.png``, regenerating
    it when the source photo has been updated. Returns None if there's
    no source photo on disk (in that case the cell is just left blank)."""
    src = src_dir / f"{name}.png"
    if not src.exists():
        return None
    THUMB_DIR.mkdir(exist_ok=True)
    dst = THUMB_DIR / f"{src_dir.name}__{name}_thumb.png"
    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        return dst
    img = PILImage.open(src)
    img.thumbnail((10_000, THUMB_HEIGHT_PX))
    img.save(dst, "PNG")
    return dst


def _write_row_by_header(ws: Any, row: int, values: Dict[str, Any]) -> None:
    """Write ``values`` to ``row`` using the sheet's actual header row to
    decide which column each key goes in. Keys that aren't present in the
    header row are silently ignored; columns that aren't in ``values``
    are left blank. This is what keeps us robust against the file having
    extra/older columns we don't know about."""
    headers = _sheet_headers(ws)
    for key, val in values.items():
        if key in headers:
            col = headers.index(key) + 1
            ws.cell(row=row, column=col, value=val)


def _write_row_by_header(ws: Any, row: int,
                         headers: List[str],
                         values: Dict[str, Any]) -> None:
    """Write ``values`` to ``row`` using ``headers`` (read from the
    sheet's header row) to decide which column each key goes in. Keys
    that aren't in the header row are silently ignored; columns that
    aren't in ``values`` are left blank."""
    for key, val in values.items():
        if key in headers:
            col = headers.index(key) + 1
            ws.cell(row=row, column=col, value=val)


def append_trial(session_id: str, t: TrialResult, comment: str = "") -> None:
    """Append one trial as a single row to ``beer-pour-results.xlsx``. Empty rows
    are not written; every call adds exactly one new row regardless of
    how many windows were in the trial. The tag photo (when present in
    ``images/tags/<tag>.png``) is embedded in the row's ``tag_photo``
    column."""
    wb = load_workbook(RESULTS_XLSX)
    trials = wb["Trials"]
    trial_headers = _sheet_headers(trials)

    next_row = trials.max_row + 1

    trial_row: Dict[str, Any] = {
        "session_id":           session_id,
        "scenario":             t.scenario,
        "power_mw":             t.power_mw,
        "scans_in_3s":          t.scans_in_3s,
        "winning_antenna":      t.winning_antenna_per_tag,
        "cross_reads":          t.cross_reads,
        "best_rssi_winner_dbm": t.best_rssi_winner
                                if t.best_rssi_winner is not None else "",
        "detected_epcs":        ", ".join(t.detected_epcs),
        "tag_photo":            "",
        "result":               t.result,
        "notes":                comment,
    }
    _write_row_by_header(trials, next_row, trial_headers, trial_row)
    trials.row_dimensions[next_row].height = ROW_HEIGHT_PT

    result_col = (trial_headers.index("result") + 1
                  if "result" in trial_headers else None)
    _style_data_row(trials, next_row, trial_headers,
                    _CENTERED_TRIALS,
                    result_col=result_col, result_value=t.result)

    # Embed the tag photo thumbnail in the row's tag_photo column.
    if "tag_photo" in trial_headers:
        thumb = _thumb_for(TAGS_DIR, t.tag)
        if thumb is not None:
            col_letter = get_column_letter(trial_headers.index("tag_photo") + 1)
            img = XLImage(str(thumb))
            img.anchor = f"{col_letter}{next_row}"
            trials.add_image(img)

    wb.save(RESULTS_XLSX)


# ────────────────────── rfid_gc_live subprocess ─────────────────────


def discover_tags() -> List[str]:
    """Tag types are whatever PNGs live in images/tags/, sorted. Adding
    a new tag = drop ``<name>.png`` in there and restart the harness."""
    if not TAGS_DIR.exists():
        return []
    return sorted(
        p.stem for p in TAGS_DIR.glob("*.png") if not p.name.startswith(".")
    )


def run_one_trial(scenario: str,
                  power_mw: int,
                  tag: str,
                  trial_num: int) -> TrialResult:
    """Launch one ``rfid_gc_live`` subprocess for a single trial. The
    operator should be ready to slide the cup the moment the harness
    prints ``GO!``; that moment is t = 0 for the trial."""
    print(
        f"\n[Trial #{trial_num}] Launching reader at {power_mw} mW... "
        f"(reader takes ~1-2 s to come up)"
    )

    proc = subprocess.Popen(
        [str(RFID_BINARY), str(power_mw)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        bufsize=1,
        universal_newlines=True,
    )

    windows: List[WindowRead] = []
    reader_ready = threading.Event()
    start_time_holder: List[Tuple[dt.datetime, float]] = []  # (wallclock, perf_counter)

    def reader_thread() -> None:
        assert proc.stdout is not None
        win_idx = 0
        for line in proc.stdout:
            sys.stdout.write("    " + line)
            sys.stdout.flush()
            clean = ANSI_RE.sub("", line)
            if not reader_ready.is_set() and READY_RE.search(clean):
                start_time_holder.append((dt.datetime.now(), time.perf_counter()))
                reader_ready.set()
                print(
                    f"[Trial #{trial_num}] GO! Slide the {tag} cup over "
                    f"either antenna now ({TRIAL_DURATION_S:.1f} s)."
                )
                sys.stdout.flush()
                continue
            if not reader_ready.is_set():
                continue
            t_offset = time.perf_counter() - start_time_holder[0][1]
            win_idx += 1
            w = parse_window_line(win_idx, t_offset, line)
            if w is None:
                win_idx -= 1
                continue
            windows.append(w)

    t = threading.Thread(target=reader_thread, daemon=True)
    t.start()

    if not reader_ready.wait(timeout=15.0):
        print(f"[Trial #{trial_num}] ERROR: reader never became ready; "
              f"check USB / power and try again.")
        _shutdown(proc)
        t.join(timeout=2)
        return TrialResult(
            scenario=scenario, power_mw=power_mw, tag=tag,
            trial_num=trial_num,
            start_time=dt.datetime.now(), duration_s=0.0,
        )

    time.sleep(TRIAL_DURATION_S)
    duration_s = time.perf_counter() - start_time_holder[0][1]

    _shutdown(proc)
    t.join(timeout=2)

    return TrialResult(
        scenario=scenario,
        power_mw=power_mw,
        tag=tag,
        trial_num=trial_num,
        start_time=start_time_holder[0][0],
        duration_s=duration_s,
        windows=windows,
    )


def _shutdown(proc: subprocess.Popen) -> None:
    """Send SIGINT so the C binary's signal handler runs
    CAENRFID_Disconnect() cleanly; escalate if it ignores us."""
    try:
        proc.send_signal(signal.SIGINT)
        proc.wait(timeout=3)
    except Exception:
        try:
            proc.terminate()
            proc.wait(timeout=2)
        except Exception:
            proc.kill()


# ───────────────────────────── menu loop ─────────────────────────────


def _pick(label: str, options: List[str]) -> int:
    while True:
        print(f"\nAvailable {label}:")
        for i, s in enumerate(options, 1):
            print(f"  {i}. {s}")
        choice = input(f"Select {label} [1..{len(options)}]: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(options):
            return int(choice) - 1
        print(f"  Invalid choice, try again.")


def _prompt_save() -> bool:
    """Ask whether to keep this trial. ENTER (or 'y'/'yes') keeps it;
    'n'/'no' discards. Repeats on anything else; Ctrl-C / EOF discards."""
    while True:
        try:
            ans = input("    Save this trial? [Y/n]: ").strip().lower()
        except (KeyboardInterrupt, EOFError):
            return False
        if ans in {"", "y", "yes"}:
            return True
        if ans in {"n", "no"}:
            return False
        print("    Type 'y' (or ENTER) to save, 'n' to discard.")


def _prompt_comment() -> str:
    """Optional free-text comment written to the trial's ``notes`` column.
    ENTER alone = no comment. Anything else is taken verbatim. Ctrl-C /
    EOF also yields no comment (so a runaway keypress can't abort the
    save we already confirmed)."""
    try:
        text = input("    Comment for notes column (ENTER to skip): ")
    except (KeyboardInterrupt, EOFError):
        return ""
    return text.strip()


def main() -> int:
    if not RFID_BINARY.is_file() or not os.access(RFID_BINARY, os.X_OK):
        print(f"ERROR: '{RFID_BINARY.name}' not found or not executable.")
        print("Build it first:  ./compile_gc.sh   (or ./run_test.sh which does it for you)")
        return 1

    tags_available = discover_tags()
    if not tags_available:
        print(f"ERROR: no tag photos found in {TAGS_DIR}.")
        print("Drop a PNG named e.g. 'foam.png' in there to register a tag type.")
        return 1

    ensure_workbook()

    session_id = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    print(f"=== Beer-pour RFID verification test session {session_id} ===")
    print(f"Results : {RESULTS_XLSX}")
    print(f"Reader  : {RFID_BINARY.name}")
    print(f"Trial   : {TRIAL_DURATION_S:.1f} s  (verify deadline: "
          f"{VERIFY_DEADLINE_S:.1f} s)")

    scenario_idx = _pick("scenarios", SCENARIOS)
    power_idx    = _pick("power levels (mW)", [str(p) for p in POWER_LEVELS_MW])
    tag_idx      = _pick("tag types", tags_available)

    # Per-(scenario, power, tag) trial counters so each unique
    # combination has its own 1, 2, 3 sequence in the spreadsheet.
    counters: Dict[Tuple[str, int, str], int] = {}

    try:
        while True:
            scenario = SCENARIOS[scenario_idx]
            power_mw = POWER_LEVELS_MW[power_idx]
            tag      = tags_available[tag_idx]
            key      = (scenario, power_mw, tag)

            print(
                f"\n[{scenario} | {power_mw} mW | {tag}]"
            )
            print(
                "  ENTER = start trial   'p' = power   's' = scenario   "
                "'t' = tag   'q' = quit"
            )
            try:
                choice = input("> ").strip().lower()
            except (KeyboardInterrupt, EOFError):
                print("\nBye.")
                return 0

            if choice == "q":
                print("Bye.")
                return 0
            if choice == "p":
                power_idx = _pick("power levels (mW)",
                                  [str(p) for p in POWER_LEVELS_MW])
                continue
            if choice == "s":
                scenario_idx = _pick("scenarios", SCENARIOS)
                continue
            if choice == "t":
                tag_idx = _pick("tag types", tags_available)
                continue
            if choice != "":
                # Anything other than ENTER / known shortcut: re-prompt.
                continue

            tentative_num = counters.get(key, 0) + 1
            result = run_one_trial(
                scenario=scenario,
                power_mw=power_mw,
                tag=tag,
                trial_num=tentative_num,
            )

            # Single-line verdict, ASCII-only so it renders the same in
            # all terminals / log files.
            if result.verified:
                rssi = (f"{result.best_rssi_winner:.1f} dBm"
                        if result.best_rssi_winner is not None else "?")
                homes = result.winning_antenna_per_tag or "(none)"
                cross_msg = (
                    f"{result.cross_reads} leaking EPC(s): "
                    f"{', '.join(result.cross_read_epcs)}"
                    if result.cross_read_epcs
                    else "0 leaking EPC(s)"
                )
                print(
                    f"[Trial #{tentative_num}] {result.result} -- "
                    f"verified in {result.ttv_s:.2f} s, "
                    f"homes [{homes}], "
                    f"{result.scans_in_3s} scans in 3 s "
                    f"(ant0={result.scans_ant0_in_3s}, "
                    f"ant1={result.scans_ant1_in_3s}), "
                    f"{cross_msg}, "
                    f"best RSSI {rssi}"
                )
            else:
                print(
                    f"[Trial #{tentative_num}] {result.result} -- "
                    f"no attribution in {result.n_windows} window(s)"
                )

            if not _prompt_save():
                print("    discarded (not saved).")
                continue

            comment = _prompt_comment()

            # Confirmed save -> commit the per-key counter and append.
            counters[key] = tentative_num
            try:
                append_trial(session_id, result, comment=comment)
                print(f"    logged to {RESULTS_XLSX.name}")
            except Exception as exc:
                print(f"    ERROR writing to {RESULTS_XLSX.name}: {exc}")
                print("    (trial was NOT saved -- fix the issue and retry)")
                # Roll the counter back so the next save reuses this num.
                counters[key] = tentative_num - 1
    except KeyboardInterrupt:
        print("\nBye.")
        return 0


if __name__ == "__main__":
    sys.exit(main())
