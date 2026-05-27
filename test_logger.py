"""Beer-pour RFID reliability test logger.

Drives the existing ``rfid_gc_live`` binary one trial at a time and
records every trial to a single master Excel workbook (``results.xlsx``)
with two sheets:

- ``Trials``  : one row per 3-second trial — scenario (tag type, layout,
  power, condition), miss-read / cross-read counts, mean scan rate, mean
  RSSI, plus an embedded thumbnail of the tag-type reference image.
- ``Windows`` : one row per 250 ms decision window — the raw per-window
  data exactly as the C binary emits it.

The binary's arbitration logic (RSSI floor, dominance, count threshold)
is preserved unchanged; this harness only orchestrates *which* power and
mug-layout each trial uses and computes the higher-level miss-read /
cross-read metrics from the windows the binary produces.

Operator workflow (on the Pi, no SSH)::

    $ ./run_test.sh
    # or:
    $ python3 test_logger.py

    === Beer-pour RFID test session 20260527-130000 ===
    Results: /home/stratela/Test-2/test/Test-2/results.xlsx

    Tag type for this session: foam (Foam tag (CP15710-01))

    Available scenarios:
       1. foam | ant0_only | 30 mW  | dry
       2. foam | ant0_only | 30 mW  | wet
       3. foam | ant0_only | 175 mW | dry
       ...
      18. foam | both      | 316 mW | wet
    Select scenario [1-18, 's' to switch tag, 'q' to quit]: 1

    [Scenario foam | ant0_only | 30 mW | dry]
        Setup : Place ONE mug centred over antenna 0. Antenna 1 must be empty.
        Liquid: Make sure the drip tray and the mug are completely DRY.

    > Press ENTER once the setup is in place ...
    [Trial #1] starting reader for 3.0 s ...
        [RFID startup messages from the C binary ...]
    [Trial #1] LIVE — recording 3.0 s of windows now.
        [GC_WIN] win=0 t_start=... ...
        [GC_WIN] win=1 t_start=... ...
        ...
    [Trial #1] DONE — 12 windows in 3.0 s
        Ant0 attribution: 91.7%   Ant1 attribution: 0.0%
        Miss-reads     : 1        Cross-reads     : 0
        Mean RSSI ant0 : -52.1 dBm
        Append this trial to results.xlsx? [Y/n] y
            logged to results.xlsx
        Another trial in this scenario? [y/N] n
"""
from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import signal
import statistics
import subprocess
import sys
import threading
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


SCRIPT_DIR    = Path(__file__).resolve().parent
RFID_BINARY   = SCRIPT_DIR / "rfid_gc_live"
IMAGES_DIR    = SCRIPT_DIR / "images"
RESULTS_XLSX  = SCRIPT_DIR / "results.xlsx"
THUMB_DIR     = SCRIPT_DIR / ".thumbs"

# Trial parameters.  Edit here, not in the C code — the binary just runs
# whatever --duration / --window-ms / power it's told to.
TRIAL_DURATION_S      = 3.0
DECISION_WINDOW_MS    = 250
POWERS_MW: List[int]  = [30, 175, 316]


# Tag types currently under test. Add another entry + drop a matching
# images/tags/<id>.png to extend the test without code changes.
@dataclass
class TagType:
    id: str
    label: str
    notes: str = ""

    @property
    def reference_image(self) -> Path:
        return IMAGES_DIR / "tags" / f"{self.id}.png"


TAG_TYPES: List[TagType] = [
    TagType(
        id="foam",
        label="Foam tag (CP15710-01)",
        notes="White-foam-backed RFID inlay, mounted on the bottom of the mug.",
    ),
    # Example future entry — uncomment + drop images/tags/sticker.png
    # TagType(id="sticker", label="Generic UHF sticker", notes="Paper inlay."),
]


# Mug layouts. expected_ant0/1 tells the analyser which antennas SHOULD
# see a tag this trial; an attribution on a "not expected" antenna is a
# cross-read, an empty slot on an "expected" antenna is a miss-read.
@dataclass
class Layout:
    id: str
    label: str
    expected_ant0: bool
    expected_ant1: bool
    operator_prompt: str


LAYOUTS: List[Layout] = [
    Layout(
        id="ant0_only",
        label="Mug over antenna 0 only",
        expected_ant0=True,
        expected_ant1=False,
        operator_prompt="Place ONE mug centred over antenna 0. Antenna 1 must be empty.",
    ),
    Layout(
        id="ant1_only",
        label="Mug over antenna 1 only",
        expected_ant0=False,
        expected_ant1=True,
        operator_prompt="Place ONE mug centred over antenna 1. Antenna 0 must be empty.",
    ),
    Layout(
        id="both",
        label="One mug over each antenna",
        expected_ant0=True,
        expected_ant1=True,
        operator_prompt="Place ONE mug centred over each antenna. Both antennas have a mug.",
    ),
]


@dataclass
class Condition:
    id: str
    label: str
    operator_prompt: str


CONDITIONS: List[Condition] = [
    Condition(
        id="dry",
        label="Dry (drip tray empty, mug empty)",
        operator_prompt="Make sure the drip tray and the mug are completely DRY. No droplets.",
    ),
    Condition(
        id="wet",
        label="Wet (beer pouring into mug)",
        operator_prompt=(
            "Start the beer pour the moment you press ENTER. Aim into the mug; "
            "some drips into the tray are expected and part of the test."
        ),
    ),
]


# Reliability thresholds. Cells of the Summary sheet that violate these
# are filled red so you can scan a session at a glance.
ATTRIBUTION_RATE_MIN_PCT = 80.0
CROSS_READ_RATE_MAX_PCT  = 0.0


# rfid_gc_live emits one structured line per window:
#   [GC_WIN] win=0 t_start=... t_end=... power_mw=30 win_ms=250.1
#            scans_0=34 scans_1=33 rate_0=135.9 rate_1=131.8
#            ant0_epc=E280...    ant0_rssi=-58.3
#            ant1_epc=-          ant1_rssi=-
#            unique=1 drop_floor=0 drop_low=2 drop_amb=0
ANSI_RE         = re.compile(r"\x1b\[[0-9;]*m")
WIN_LINE_RE     = re.compile(r"\[GC_WIN\]\s+(?P<kv>.+)")
KV_RE           = re.compile(r"(\w+)=(\S+)")
# Printed once by the C binary right before it enters the polling loop.
# Used to detect "the reader is now actually scanning, recording is live".
READER_READY_RE = re.compile(r"\[GC\] Ready\. One line every")


TRIAL_HEADERS = [
    "session_id",
    "trial_uid",
    "started_at",
    "tag_type",
    "layout",
    "power_mw",
    "condition",
    "expected_ant0",
    "expected_ant1",
    "total_windows",
    "ant0_attribution_%",
    "ant1_attribution_%",
    "ant0_inferred_epc",
    "ant1_inferred_epc",
    "miss_read_count",
    "cross_read_count",
    "mean_rate_0_hz",
    "mean_rate_1_hz",
    "mean_rssi_ant0_dbm",
    "mean_rssi_ant1_dbm",
    "drop_floor_total",
    "drop_low_total",
    "drop_amb_total",
    "tag_reference_photo",
    "notes",
]

WINDOW_HEADERS = [
    "trial_uid",
    "win",
    "t_start_unix",
    "t_end_unix",
    "power_mw",
    "win_ms",
    "scans_0",
    "scans_1",
    "rate_0",
    "rate_1",
    "ant0_epc",
    "ant0_rssi",
    "ant1_epc",
    "ant1_rssi",
    "unique",
    "drop_floor",
    "drop_low",
    "drop_amb",
]

TRIAL_WIDTHS  = [16, 38, 20, 12, 12, 9, 10, 14, 14, 13, 18, 18,
                 22, 22, 14, 14, 14, 14, 18, 18, 14, 14, 14, 24, 24]
WINDOW_WIDTHS = [38, 5, 16, 16, 9, 8, 8, 8, 8, 8,
                 26, 18, 26, 18, 8, 11, 9, 9]

THUMB_HEIGHT_PX = 80
ROW_HEIGHT_PT   = 64

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
FAIL_FILL   = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


# ─────────────────────────── data types ────────────────────────────


@dataclass
class WindowRow:
    """One [GC_WIN] line parsed into structured fields."""
    win: int
    t_start: float
    t_end: float
    power_mw: int
    win_ms: float
    scans_0: int
    scans_1: int
    rate_0: float
    rate_1: float
    ant0_epcs: List[str]
    ant0_rssis: List[float]
    ant1_epcs: List[str]
    ant1_rssis: List[float]
    unique: int
    drop_floor: int
    drop_low: int
    drop_amb: int


@dataclass
class TrialResult:
    session_id: str
    trial_uid: str
    started_at: dt.datetime
    tag: TagType
    layout: Layout
    power_mw: int
    condition: Condition
    windows: List[WindowRow] = field(default_factory=list)
    notes: str = ""

    # Derived (filled in by compute_metrics).
    ant0_attribution_pct: float = 0.0
    ant1_attribution_pct: float = 0.0
    ant0_inferred_epc: str = ""
    ant1_inferred_epc: str = ""
    miss_read_count: int = 0
    cross_read_count: int = 0
    mean_rate_0_hz: float = 0.0
    mean_rate_1_hz: float = 0.0
    mean_rssi_ant0_dbm: Optional[float] = None
    mean_rssi_ant1_dbm: Optional[float] = None
    drop_floor_total: int = 0
    drop_low_total: int = 0
    drop_amb_total: int = 0


# ─────────────────────────── parsing ────────────────────────────


def _parse_pipe_strings(s: str) -> List[str]:
    if not s or s == "-":
        return []
    return [tok for tok in s.split("|") if tok]


def _parse_pipe_floats(s: str) -> List[float]:
    return [float(x) for x in _parse_pipe_strings(s)]


def parse_window_line(line: str) -> Optional[WindowRow]:
    """Parse one [GC_WIN] line. Returns None if the line isn't one."""
    clean = ANSI_RE.sub("", line)
    m = WIN_LINE_RE.search(clean)
    if not m:
        return None
    kv = dict(KV_RE.findall(m.group("kv")))
    try:
        return WindowRow(
            win=int(kv["win"]),
            t_start=float(kv["t_start"]),
            t_end=float(kv["t_end"]),
            power_mw=int(kv["power_mw"]),
            win_ms=float(kv["win_ms"]),
            scans_0=int(kv["scans_0"]),
            scans_1=int(kv["scans_1"]),
            rate_0=float(kv["rate_0"]),
            rate_1=float(kv["rate_1"]),
            ant0_epcs=_parse_pipe_strings(kv.get("ant0_epc", "-")),
            ant0_rssis=_parse_pipe_floats(kv.get("ant0_rssi", "-")),
            ant1_epcs=_parse_pipe_strings(kv.get("ant1_epc", "-")),
            ant1_rssis=_parse_pipe_floats(kv.get("ant1_rssi", "-")),
            unique=int(kv["unique"]),
            drop_floor=int(kv["drop_floor"]),
            drop_low=int(kv["drop_low"]),
            drop_amb=int(kv["drop_amb"]),
        )
    except (KeyError, ValueError) as exc:
        print(f"    [WARN] could not parse [GC_WIN] line: {exc}")
        return None


# ─────────────────────────── metrics ────────────────────────────


def compute_metrics(trial: TrialResult) -> None:
    """Derive the per-trial reliability numbers from `trial.windows`.

    Miss-read rule: in a window where the layout says antenna N has a
    mug above it but ant<N>_epc is empty, that's a miss-read.

    Cross-read rule:
      * For single-mug layouts (ant0_only / ant1_only), any attribution
        on the "empty" antenna is a cross-read regardless of EPC.
      * For the `both` layout, the EPC modally appearing on each antenna
        across the trial is treated as "the mug that belongs there";
        if that EPC ever appears on the OTHER antenna's slot during the
        trial, that window counts as a cross-read.
    """
    n = len(trial.windows)
    trial.windows_total = n  # type: ignore[attr-defined]
    if n == 0:
        return

    ant0_counter: Counter = Counter()
    ant1_counter: Counter = Counter()
    ant0_rssis: List[float] = []
    ant1_rssis: List[float] = []
    rate0: List[float] = []
    rate1: List[float] = []
    ant0_nonempty = 0
    ant1_nonempty = 0
    df = dl = da = 0

    for w in trial.windows:
        if w.ant0_epcs:
            ant0_nonempty += 1
            ant0_counter.update(w.ant0_epcs)
            ant0_rssis.extend(w.ant0_rssis)
        if w.ant1_epcs:
            ant1_nonempty += 1
            ant1_counter.update(w.ant1_epcs)
            ant1_rssis.extend(w.ant1_rssis)
        rate0.append(w.rate_0)
        rate1.append(w.rate_1)
        df += w.drop_floor
        dl += w.drop_low
        da += w.drop_amb

    trial.ant0_attribution_pct = 100.0 * ant0_nonempty / n
    trial.ant1_attribution_pct = 100.0 * ant1_nonempty / n
    trial.ant0_inferred_epc = ant0_counter.most_common(1)[0][0] if ant0_counter else ""
    trial.ant1_inferred_epc = ant1_counter.most_common(1)[0][0] if ant1_counter else ""
    trial.mean_rate_0_hz = statistics.fmean(rate0) if rate0 else 0.0
    trial.mean_rate_1_hz = statistics.fmean(rate1) if rate1 else 0.0
    trial.mean_rssi_ant0_dbm = statistics.fmean(ant0_rssis) if ant0_rssis else None
    trial.mean_rssi_ant1_dbm = statistics.fmean(ant1_rssis) if ant1_rssis else None
    trial.drop_floor_total = df
    trial.drop_low_total = dl
    trial.drop_amb_total = da

    miss = 0
    cross = 0
    for w in trial.windows:
        if trial.layout.expected_ant0 and not w.ant0_epcs:
            miss += 1
        if trial.layout.expected_ant1 and not w.ant1_epcs:
            miss += 1

        if not trial.layout.expected_ant0 and w.ant0_epcs:
            cross += 1
        if not trial.layout.expected_ant1 and w.ant1_epcs:
            cross += 1

        if trial.layout.expected_ant0 and trial.layout.expected_ant1:
            if trial.ant0_inferred_epc and trial.ant0_inferred_epc in w.ant1_epcs:
                cross += 1
            if trial.ant1_inferred_epc and trial.ant1_inferred_epc in w.ant0_epcs:
                cross += 1

    trial.miss_read_count = miss
    trial.cross_read_count = cross


# ─────────────────────────── workbook I/O ────────────────────────────


def ensure_workbook() -> None:
    if not RESULTS_XLSX.exists():
        wb = Workbook()
        trials = wb.active
        trials.title = "Trials"
        trials.append(TRIAL_HEADERS)
        _style_header_row(trials)
        for col, w in enumerate(TRIAL_WIDTHS, start=1):
            trials.column_dimensions[get_column_letter(col)].width = w

        windows = wb.create_sheet("Windows")
        windows.append(WINDOW_HEADERS)
        _style_header_row(windows)
        for col, w in enumerate(WINDOW_WIDTHS, start=1):
            windows.column_dimensions[get_column_letter(col)].width = w

        wb.save(RESULTS_XLSX)
        return

    wb = load_workbook(RESULTS_XLSX)
    changed = False

    def upgrade(sheet_name: str, headers: List[str], widths: List[int]) -> None:
        nonlocal changed
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            _style_header_row(ws)
            for col, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = w
            changed = True
            return
        ws = wb[sheet_name]
        for i, h in enumerate(headers, start=1):
            existing = ws.cell(row=1, column=i).value
            if existing is None or existing == "":
                ws.cell(row=1, column=i, value=h)
                changed = True
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w
        _style_header_row(ws)

    upgrade("Trials", TRIAL_HEADERS, TRIAL_WIDTHS)
    upgrade("Windows", WINDOW_HEADERS, WINDOW_WIDTHS)
    if changed:
        wb.save(RESULTS_XLSX)


def _style_header_row(ws) -> None:
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL


def thumb_for(tag: TagType) -> Optional[Path]:
    src = tag.reference_image
    if not src.exists():
        return None
    THUMB_DIR.mkdir(exist_ok=True)
    dst = THUMB_DIR / f"tag_{tag.id}_thumb.png"
    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        return dst
    img = PILImage.open(src)
    img.thumbnail((10_000, THUMB_HEIGHT_PX))
    img.convert("RGB").save(dst, "PNG")
    return dst


def append_trial(trial: TrialResult) -> None:
    """Append `trial` to results.xlsx: one row in Trials with thumbnail,
    one row per window in Windows."""
    wb = load_workbook(RESULTS_XLSX)
    trials_ws = wb["Trials"]
    windows_ws = wb["Windows"]

    photo_col_letter = get_column_letter(len(TRIAL_HEADERS) - 1)  # tag_reference_photo
    next_row = trials_ws.max_row + 1

    def fmt_rssi(x: Optional[float]) -> str:
        return f"{x:.2f}" if x is not None else ""

    trials_ws.append([
        trial.session_id,
        trial.trial_uid,
        trial.started_at.strftime("%Y-%m-%d %H:%M:%S"),
        trial.tag.id,
        trial.layout.id,
        trial.power_mw,
        trial.condition.id,
        trial.layout.expected_ant0,
        trial.layout.expected_ant1,
        len(trial.windows),
        round(trial.ant0_attribution_pct, 2),
        round(trial.ant1_attribution_pct, 2),
        trial.ant0_inferred_epc,
        trial.ant1_inferred_epc,
        trial.miss_read_count,
        trial.cross_read_count,
        round(trial.mean_rate_0_hz, 2),
        round(trial.mean_rate_1_hz, 2),
        fmt_rssi(trial.mean_rssi_ant0_dbm),
        fmt_rssi(trial.mean_rssi_ant1_dbm),
        trial.drop_floor_total,
        trial.drop_low_total,
        trial.drop_amb_total,
        "",
        trial.notes,
    ])
    trials_ws.row_dimensions[next_row].height = ROW_HEIGHT_PT

    # Flag the row red if it violated either reliability threshold.
    if trial.layout.expected_ant0 and trial.ant0_attribution_pct < ATTRIBUTION_RATE_MIN_PCT:
        trials_ws.cell(row=next_row, column=11).fill = FAIL_FILL
    if trial.layout.expected_ant1 and trial.ant1_attribution_pct < ATTRIBUTION_RATE_MIN_PCT:
        trials_ws.cell(row=next_row, column=12).fill = FAIL_FILL
    if trial.cross_read_count > 0:
        trials_ws.cell(row=next_row, column=16).fill = FAIL_FILL

    thumb = thumb_for(trial.tag)
    if thumb is not None:
        img = XLImage(str(thumb))
        img.anchor = f"{photo_col_letter}{next_row}"
        trials_ws.add_image(img)

    for w in trial.windows:
        windows_ws.append([
            trial.trial_uid,
            w.win,
            w.t_start,
            w.t_end,
            w.power_mw,
            round(w.win_ms, 2),
            w.scans_0,
            w.scans_1,
            round(w.rate_0, 2),
            round(w.rate_1, 2),
            "|".join(w.ant0_epcs),
            "|".join(f"{x:.1f}" for x in w.ant0_rssis),
            "|".join(w.ant1_epcs),
            "|".join(f"{x:.1f}" for x in w.ant1_rssis),
            w.unique,
            w.drop_floor,
            w.drop_low,
            w.drop_amb,
        ])

    wb.save(RESULTS_XLSX)


# ────────────────────── rfid_gc_live subprocess ──────────────────────


def run_one_trial(*, tag: TagType, layout: Layout, power_mw: int,
                  condition: Condition, session_id: str,
                  trial_num: int) -> TrialResult:
    """Spawn `rfid_gc_live` with --duration / --window-ms, parse its
    [GC_WIN] lines into WindowRow objects, return the TrialResult."""
    trial_uid = (
        f"{tag.id}_{layout.id}_{power_mw}mW_{condition.id}"
        f"_t{trial_num}_{dt.datetime.now().strftime('%H%M%S')}"
    )

    print(
        f"\n[Trial #{trial_num}] starting reader for {TRIAL_DURATION_S:.1f} s "
        f"(power={power_mw} mW, window={DECISION_WINDOW_MS} ms) ..."
    )
    started_at = dt.datetime.now()

    cmd = [
        str(RFID_BINARY),
        str(power_mw),
        "--duration", str(TRIAL_DURATION_S),
        "--window-ms", str(DECISION_WINDOW_MS),
    ]
    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        bufsize=1,
        universal_newlines=True,
    )

    windows: List[WindowRow] = []
    reader_ready = threading.Event()

    def reader_thread() -> None:
        assert proc.stdout is not None
        for line in proc.stdout:
            clean = ANSI_RE.sub("", line)
            if not reader_ready.is_set() and READER_READY_RE.search(clean):
                reader_ready.set()
                print(f"[Trial #{trial_num}] LIVE — recording {TRIAL_DURATION_S:.1f} s "
                      f"of windows now.")
            wr = parse_window_line(line)
            if wr is not None:
                windows.append(wr)
                # Show a brief one-line summary of the just-recorded window.
                summary_0 = wr.ant0_epcs[0] if wr.ant0_epcs else "-"
                summary_1 = wr.ant1_epcs[0] if wr.ant1_epcs else "-"
                print(f"    win={wr.win:2d}  ant0={summary_0}  ant1={summary_1}  "
                      f"drop(floor/low/amb)={wr.drop_floor}/{wr.drop_low}/{wr.drop_amb}")
            else:
                # Pass through the C banner / connect logs so the operator
                # can see them, indented to distinguish from our own output.
                sys.stdout.write("    " + line)
                sys.stdout.flush()

    t = threading.Thread(target=reader_thread, daemon=True)
    t.start()

    # The binary auto-stops after --duration. We give it a generous wait
    # cushion in case it's slow to disconnect from the reader.
    try:
        proc.wait(timeout=TRIAL_DURATION_S + 30)
    except subprocess.TimeoutExpired:
        print(f"[Trial #{trial_num}] WARN: binary did not exit on its own, "
              f"sending SIGINT ...")
        try:
            proc.send_signal(signal.SIGINT)
            proc.wait(timeout=3)
        except Exception:
            try:
                proc.terminate(); proc.wait(timeout=2)
            except Exception:
                proc.kill()

    t.join(timeout=2)

    if not reader_ready.is_set():
        print(f"[Trial #{trial_num}] WARN: reader never reached LIVE — "
              f"check that {RFID_BINARY.name} can open /dev/ttyACM0.")

    print(f"[Trial #{trial_num}] DONE — {len(windows)} window(s) in "
          f"{(dt.datetime.now() - started_at).total_seconds():.1f} s")

    return TrialResult(
        session_id=session_id,
        trial_uid=trial_uid,
        started_at=started_at,
        tag=tag,
        layout=layout,
        power_mw=power_mw,
        condition=condition,
        windows=windows,
    )


# ───────────────────────────── menu loop ─────────────────────────────


def all_scenarios(tag: TagType) -> List[Tuple[Layout, int, Condition]]:
    """Cartesian product of layouts × powers × conditions for `tag`."""
    out: List[Tuple[Layout, int, Condition]] = []
    for layout in LAYOUTS:
        for power in POWERS_MW:
            for cond in CONDITIONS:
                out.append((layout, power, cond))
    return out


def pick_tag_type() -> TagType:
    if len(TAG_TYPES) == 1:
        print(f"\nTag type for this session: {TAG_TYPES[0].id} ({TAG_TYPES[0].label})")
        return TAG_TYPES[0]
    while True:
        print("\nAvailable tag types:")
        for i, t in enumerate(TAG_TYPES, 1):
            print(f"  {i}. {t.id}  ({t.label})")
        raw = input(f"Select tag type [1-{len(TAG_TYPES)}]: ").strip()
        if raw.isdigit():
            i = int(raw)
            if 1 <= i <= len(TAG_TYPES):
                return TAG_TYPES[i - 1]
        print("  invalid choice, try again.")


def pick_scenario(tag: TagType) -> Optional[Tuple[Layout, int, Condition]]:
    """Returns the chosen (layout, power, condition), or None for 's'/'q'.

    On 'q' the function calls sys.exit(0); on 's' it returns None so the
    caller can re-pick the tag type.
    """
    scenarios = all_scenarios(tag)
    while True:
        print(f"\nAvailable scenarios for tag '{tag.id}':")
        for i, (layout, power, cond) in enumerate(scenarios, 1):
            print(f"  {i:2d}. {tag.id} | {layout.id:9s} | {power:3d} mW | {cond.id}")
        raw = input(
            f"Select scenario [1-{len(scenarios)}, 's' to switch tag, 'q' to quit]: "
        ).strip().lower()
        if raw == "q":
            print("Bye.")
            sys.exit(0)
        if raw == "s":
            return None
        if raw.isdigit():
            i = int(raw)
            if 1 <= i <= len(scenarios):
                return scenarios[i - 1]
        print("  invalid choice, try again.")


def yes_no(question: str, default: bool) -> bool:
    suffix = " [Y/n] " if default else " [y/N] "
    while True:
        try:
            raw = input(question + suffix).strip().lower()
        except (KeyboardInterrupt, EOFError):
            return False
        if not raw:
            return default
        if raw in ("y", "yes"):
            return True
        if raw in ("n", "no"):
            return False
        print("    please answer y or n.")


def print_trial_summary(t: TrialResult) -> None:
    n = len(t.windows)
    print(f"    Total windows   : {n}")
    print(f"    Ant0 attribution: {t.ant0_attribution_pct:6.2f}%   "
          f"inferred EPC: {t.ant0_inferred_epc or '-'}")
    print(f"    Ant1 attribution: {t.ant1_attribution_pct:6.2f}%   "
          f"inferred EPC: {t.ant1_inferred_epc or '-'}")
    print(f"    Miss-reads      : {t.miss_read_count}")
    print(f"    Cross-reads     : {t.cross_read_count}")
    print(f"    Scan rate 0/1   : {t.mean_rate_0_hz:.1f} Hz / {t.mean_rate_1_hz:.1f} Hz")
    r0 = f"{t.mean_rssi_ant0_dbm:.1f}" if t.mean_rssi_ant0_dbm is not None else "-"
    r1 = f"{t.mean_rssi_ant1_dbm:.1f}" if t.mean_rssi_ant1_dbm is not None else "-"
    print(f"    Mean RSSI 0/1   : {r0} dBm / {r1} dBm")
    print(f"    Dropped         : floor={t.drop_floor_total}  "
          f"low_count={t.drop_low_total}  ambiguous={t.drop_amb_total}")


def main() -> int:
    if not RFID_BINARY.is_file() or not os.access(RFID_BINARY, os.X_OK):
        print(f"ERROR: '{RFID_BINARY}' not found or not executable.")
        print("Build it first:  ./compile_gc.sh")
        return 1

    ensure_workbook()
    session_id = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    print(f"\n=== Beer-pour RFID test session {session_id} ===")
    print(f"Results: {RESULTS_XLSX}")
    print(f"Trial duration: {TRIAL_DURATION_S:.1f} s    "
          f"Decision window: {DECISION_WINDOW_MS} ms    "
          f"Powers: {', '.join(str(p)+' mW' for p in POWERS_MW)}")

    trial_counter = 0
    while True:
        tag = pick_tag_type()
        if not tag.reference_image.exists():
            print(f"  NOTE: reference image not found at {tag.reference_image}")
            print(f"        the trial will still record; the photo column will be blank.")

        while True:
            picked = pick_scenario(tag)
            if picked is None:
                break  # switch tag type
            layout, power, condition = picked

            print()
            print("-" * 72)
            print(f"  Scenario: {tag.id} | {layout.id} | {power} mW | {condition.id}")
            print(f"  Setup   : {layout.operator_prompt}")
            print(f"  Liquid  : {condition.operator_prompt}")
            print("-" * 72)

            while True:
                try:
                    input("\n  > Press ENTER once the setup is in place (Ctrl-C to back out) ...")
                except (KeyboardInterrupt, EOFError):
                    print("\n  back to the scenario menu.")
                    break

                trial_counter += 1
                try:
                    trial = run_one_trial(
                        tag=tag,
                        layout=layout,
                        power_mw=power,
                        condition=condition,
                        session_id=session_id,
                        trial_num=trial_counter,
                    )
                except KeyboardInterrupt:
                    print("\n  trial interrupted; results not saved.")
                    continue

                compute_metrics(trial)
                print()
                print_trial_summary(trial)

                if yes_no("\n  Append this trial to results.xlsx?", default=True):
                    try:
                        append_trial(trial)
                        print(f"    logged to {RESULTS_XLSX.name}")
                    except Exception as exc:
                        print(f"    ERROR writing to {RESULTS_XLSX.name}: {exc}")
                        print("    (trial was NOT saved — fix the issue and retry)")
                else:
                    print("    discarded.")

                if not yes_no("  Run another trial in THIS scenario?", default=False):
                    break

    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nbye.")
        sys.exit(0)
