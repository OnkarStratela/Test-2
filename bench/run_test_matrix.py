"""
run_test_matrix.py — interactive harness for the beer-pour RFID test.

Runs the modified `rfid_gc_live` binary on a Linux box (over SSH) with
preset combinations of:

  * tag type    (foam, ...)             — pick which tag type this session is for
  * mug layout  (ant0_only / ant1_only / both)
  * power_mw    (30, 175, 316 by default)
  * condition   (dry / wet)

For each scenario the harness:
  1. Prints the operator setup instructions ("place mug A over ant 0, dry").
  2. Waits for ENTER (so the operator can physically set things up).
  3. Runs the binary remotely with --duration 3 --window-ms 250 --csv --quiet.
  4. Pulls the CSV back to a session folder.
  5. Computes per-trial reliability (attribution % per antenna, cross-read %, mean RSSI).
  6. Asks: append this trial to the report? (y/n)
  7. Asks: another trial in this cell, or move on? (a/m)

At the end (or at any quit) it writes / updates a single
`bench/results/report.xlsx` with:

  * Summary   — one row per (tag_type, layout, power, condition) cell
  * Trials    — one row per individual trial
  * Raw       — one row per decision window across every trial
  * Tags      — reference image + notes per tag type
  * Config    — session metadata (SSH host, code hash, timestamps)

Run from the bench/ folder on Windows:

    python run_test_matrix.py [--config config.yaml] [--tag-type foam]

Dependencies: see requirements.txt.
"""
from __future__ import annotations

import argparse
import csv
import io
import os
import posixpath
import shlex
import statistics
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import paramiko
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ----------------------------------------------------------------------
# Paths.
# ----------------------------------------------------------------------
BENCH_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BENCH_DIR / "results"
REPORT_XLSX = RESULTS_DIR / "report.xlsx"

# ----------------------------------------------------------------------
# Config dataclasses (mirror config.yaml).
# ----------------------------------------------------------------------
@dataclass
class SSHConfig:
    host: str
    user: str
    port: int = 22
    key_path: str = ""
    password: str = ""
    remote_workdir: str = "~/Test-2"
    remote_binary: str = "./rfid_gc_live"
    use_sudo: bool = False


@dataclass
class MeasurementConfig:
    trial_duration_s: float = 3.0
    decision_window_ms: int = 250
    inter_trial_pause_s: float = 1.0


@dataclass
class Condition:
    id: str
    label: str
    operator_prompt: str


@dataclass
class Layout:
    id: str
    label: str
    expected: dict  # {"ant0": True, "ant1": False}
    operator_prompt: str


@dataclass
class TagType:
    id: str
    label: str
    reference_image: str = ""
    notes: str = ""


@dataclass
class Thresholds:
    attribution_rate_min_pct: float = 80.0
    cross_read_rate_max_pct: float = 0.0


@dataclass
class Config:
    ssh: SSHConfig
    measurement: MeasurementConfig
    powers_mw: list[int]
    conditions: list[Condition]
    layouts: list[Layout]
    tag_types: list[TagType]
    thresholds: Thresholds


def load_config(path: Path) -> Config:
    with path.open("r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)
    return Config(
        ssh=SSHConfig(**raw["ssh"]),
        measurement=MeasurementConfig(**raw.get("measurement", {})),
        powers_mw=list(raw["powers_mw"]),
        conditions=[Condition(**c) for c in raw["conditions"]],
        layouts=[Layout(**lay) for lay in raw["layouts"]],
        tag_types=[TagType(**t) for t in raw["tag_types"]],
        thresholds=Thresholds(**raw.get("thresholds", {})),
    )


# ----------------------------------------------------------------------
# Trial result records (what we append to the workbook).
# ----------------------------------------------------------------------
@dataclass
class WindowRow:
    """One decision-window row as parsed from the CSV."""
    window_idx: int
    t_start_unix: float
    t_end_unix: float
    power_mw: int
    window_ms: float
    scans_0: int
    scans_1: int
    rate_0_hz: float
    rate_1_hz: float
    attr_ant0_epcs: list[str]
    attr_ant0_rssis: list[float]
    attr_ant1_epcs: list[str]
    attr_ant1_rssis: list[float]
    n_unique_epcs: int
    n_dropped_below_floor: int
    n_dropped_low_count: int
    n_dropped_ambiguous: int


@dataclass
class TrialResult:
    """Reliability stats for one 3-second trial."""
    session_id: str
    trial_uid: str
    started_at_iso: str
    tag_type: str
    layout: str
    power_mw: int
    condition: str
    expected_ant0: bool
    expected_ant1: bool
    csv_path: str
    windows: list[WindowRow] = field(default_factory=list)

    # Derived metrics, computed once after CSV ingestion.
    total_windows: int = 0
    ant0_attribution_pct: float = 0.0   # % of windows where ant0 slot is non-empty
    ant1_attribution_pct: float = 0.0
    ant0_inferred_epc: str = ""         # most common EPC on ant0 across the trial
    ant1_inferred_epc: str = ""
    cross_read_count: int = 0           # windows where ant0_epc appears on ant1 or vice versa
    miss_read_count: int = 0            # windows where an expected antenna is empty
    mean_rate_0_hz: float = 0.0
    mean_rate_1_hz: float = 0.0
    mean_rssi_ant0_dbm: float = float("nan")
    mean_rssi_ant1_dbm: float = float("nan")
    total_dropped_below_floor: int = 0
    total_dropped_low_count: int = 0
    total_dropped_ambiguous: int = 0
    notes: str = ""


# ----------------------------------------------------------------------
# CSV parsing.
# ----------------------------------------------------------------------
def _split_semi(s: str) -> list[str]:
    s = (s or "").strip()
    if not s:
        return []
    return [tok for tok in s.split(";") if tok]


def parse_csv(path: Path) -> list[WindowRow]:
    rows: list[WindowRow] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(WindowRow(
                window_idx=int(r["window_idx"]),
                t_start_unix=float(r["t_start_unix"]),
                t_end_unix=float(r["t_end_unix"]),
                power_mw=int(r["power_mw"]),
                window_ms=float(r["window_ms"]),
                scans_0=int(r["scans_0"]),
                scans_1=int(r["scans_1"]),
                rate_0_hz=float(r["rate_0_hz"]),
                rate_1_hz=float(r["rate_1_hz"]),
                attr_ant0_epcs=_split_semi(r["attr_ant0_epc"]),
                attr_ant0_rssis=[float(x) for x in _split_semi(r["attr_ant0_max_rssi_dbm"])],
                attr_ant1_epcs=_split_semi(r["attr_ant1_epc"]),
                attr_ant1_rssis=[float(x) for x in _split_semi(r["attr_ant1_max_rssi_dbm"])],
                n_unique_epcs=int(r["n_unique_epcs"]),
                n_dropped_below_floor=int(r["n_dropped_below_floor"]),
                n_dropped_low_count=int(r["n_dropped_low_count"]),
                n_dropped_ambiguous=int(r["n_dropped_ambiguous"]),
            ))
    return rows


# ----------------------------------------------------------------------
# Reliability computation.
#
# Strategy chosen by the user: don't pre-declare which EPC belongs to
# which antenna. Instead, infer it per trial — the modal EPC on each
# antenna across the trial is treated as "the mug that was placed
# there". A cross-read is then any window where:
#
#   - ant0_inferred_epc appears in ant1's slot, OR
#   - ant1_inferred_epc appears in ant0's slot.
#
# A miss-read is any window where the layout says an antenna SHOULD see
# a mug but the slot is empty.
#
# For layouts with one empty antenna (ant0_only / ant1_only), the
# expected_ant of the empty side is False; any attribution there is
# counted as a cross-read (a false positive) regardless of EPC.
# ----------------------------------------------------------------------
def compute_metrics(trial: TrialResult) -> None:
    trial.total_windows = len(trial.windows)
    if trial.total_windows == 0:
        return

    ant0_epc_counter = Counter()
    ant1_epc_counter = Counter()
    ant0_rssis: list[float] = []
    ant1_rssis: list[float] = []
    rate0: list[float] = []
    rate1: list[float] = []
    ant0_nonempty = 0
    ant1_nonempty = 0
    total_drop_floor = 0
    total_drop_low = 0
    total_drop_ambig = 0

    for w in trial.windows:
        if w.attr_ant0_epcs:
            ant0_nonempty += 1
            ant0_epc_counter.update(w.attr_ant0_epcs)
            ant0_rssis.extend(w.attr_ant0_rssis)
        if w.attr_ant1_epcs:
            ant1_nonempty += 1
            ant1_epc_counter.update(w.attr_ant1_epcs)
            ant1_rssis.extend(w.attr_ant1_rssis)
        rate0.append(w.rate_0_hz)
        rate1.append(w.rate_1_hz)
        total_drop_floor += w.n_dropped_below_floor
        total_drop_low += w.n_dropped_low_count
        total_drop_ambig += w.n_dropped_ambiguous

    n = trial.total_windows
    trial.ant0_attribution_pct = 100.0 * ant0_nonempty / n
    trial.ant1_attribution_pct = 100.0 * ant1_nonempty / n
    trial.ant0_inferred_epc = ant0_epc_counter.most_common(1)[0][0] if ant0_epc_counter else ""
    trial.ant1_inferred_epc = ant1_epc_counter.most_common(1)[0][0] if ant1_epc_counter else ""
    trial.mean_rate_0_hz = statistics.fmean(rate0) if rate0 else 0.0
    trial.mean_rate_1_hz = statistics.fmean(rate1) if rate1 else 0.0
    trial.mean_rssi_ant0_dbm = statistics.fmean(ant0_rssis) if ant0_rssis else float("nan")
    trial.mean_rssi_ant1_dbm = statistics.fmean(ant1_rssis) if ant1_rssis else float("nan")
    trial.total_dropped_below_floor = total_drop_floor
    trial.total_dropped_low_count = total_drop_low
    trial.total_dropped_ambiguous = total_drop_ambig

    cross = 0
    miss = 0
    for w in trial.windows:
        if trial.expected_ant0 and not w.attr_ant0_epcs:
            miss += 1
        if trial.expected_ant1 and not w.attr_ant1_epcs:
            miss += 1

        if not trial.expected_ant0 and w.attr_ant0_epcs:
            cross += 1
        if not trial.expected_ant1 and w.attr_ant1_epcs:
            cross += 1

        if trial.expected_ant0 and trial.expected_ant1:
            if trial.ant0_inferred_epc and trial.ant0_inferred_epc in w.attr_ant1_epcs:
                cross += 1
            if trial.ant1_inferred_epc and trial.ant1_inferred_epc in w.attr_ant0_epcs:
                cross += 1

    trial.cross_read_count = cross
    trial.miss_read_count = miss


# ----------------------------------------------------------------------
# SSH execution.
# ----------------------------------------------------------------------
class RemoteRunner:
    def __init__(self, cfg: SSHConfig):
        self.cfg = cfg
        self.client: paramiko.SSHClient | None = None

    def connect(self) -> None:
        c = paramiko.SSHClient()
        c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        kwargs = {
            "hostname": self.cfg.host,
            "port": self.cfg.port,
            "username": self.cfg.user,
            "timeout": 15,
        }
        if self.cfg.key_path:
            kwargs["key_filename"] = self.cfg.key_path
        elif self.cfg.password:
            kwargs["password"] = self.cfg.password
        c.connect(**kwargs)
        self.client = c
        print(f"  [ssh] connected to {self.cfg.user}@{self.cfg.host}")

    def close(self) -> None:
        if self.client:
            self.client.close()
            self.client = None

    def run_trial(self, *, power_mw: int, duration_s: float,
                  window_ms: int, remote_csv_path: str) -> tuple[int, str, str]:
        """Run one trial. Returns (exit_status, stdout, stderr)."""
        assert self.client is not None
        c = self.cfg

        bin_cmd = (
            f"cd {shlex.quote(c.remote_workdir)} && "
            f"{shlex.quote(c.remote_binary)} {power_mw} "
            f"--duration {duration_s} "
            f"--window-ms {window_ms} "
            f"--csv {shlex.quote(remote_csv_path)} "
            f"--quiet"
        )
        if c.use_sudo:
            bin_cmd = f"sudo -n {bin_cmd}"

        stdin, stdout, stderr = self.client.exec_command(bin_cmd, timeout=duration_s + 30)
        out = stdout.read().decode("utf-8", errors="replace")
        err = stderr.read().decode("utf-8", errors="replace")
        status = stdout.channel.recv_exit_status()
        return status, out, err

    def fetch(self, remote_path: str, local_path: Path) -> None:
        assert self.client is not None
        sftp = self.client.open_sftp()
        try:
            sftp.get(remote_path, str(local_path))
            sftp.remove(remote_path)
        finally:
            sftp.close()

    def expand_remote_workdir(self) -> str:
        """Resolve ~ on the remote side once so SFTP can use an absolute path."""
        assert self.client is not None
        stdin, stdout, stderr = self.client.exec_command(
            f"cd {shlex.quote(self.cfg.remote_workdir)} && pwd",
            timeout=10,
        )
        out = stdout.read().decode("utf-8").strip()
        if not out:
            raise RuntimeError(
                f"Cannot resolve remote_workdir '{self.cfg.remote_workdir}' "
                f"on {self.cfg.host}. stderr={stderr.read().decode('utf-8', 'replace')}"
            )
        return out


# ----------------------------------------------------------------------
# Excel report (load-or-create + append rows + embed images).
# ----------------------------------------------------------------------
SUMMARY_HEADERS = [
    "tag_type", "layout", "power_mw", "condition",
    "n_trials", "n_windows",
    "ant0_attribution_%", "ant1_attribution_%",
    "cross_read_%", "miss_read_%",
    "mean_rate_0_hz", "mean_rate_1_hz",
    "mean_rssi_ant0_dbm", "mean_rssi_ant1_dbm",
    "setup_photo",
]

TRIAL_HEADERS = [
    "session_id", "trial_uid", "started_at_iso",
    "tag_type", "layout", "power_mw", "condition",
    "expected_ant0", "expected_ant1",
    "total_windows",
    "ant0_attribution_%", "ant1_attribution_%",
    "ant0_inferred_epc", "ant1_inferred_epc",
    "cross_read_count", "miss_read_count",
    "mean_rate_0_hz", "mean_rate_1_hz",
    "mean_rssi_ant0_dbm", "mean_rssi_ant1_dbm",
    "dropped_below_floor", "dropped_low_count", "dropped_ambiguous",
    "csv_path", "notes",
]

RAW_HEADERS = [
    "trial_uid", "window_idx",
    "t_start_unix", "t_end_unix", "power_mw", "window_ms",
    "scans_0", "scans_1", "rate_0_hz", "rate_1_hz",
    "attr_ant0_epc", "attr_ant0_max_rssi_dbm",
    "attr_ant1_epc", "attr_ant1_max_rssi_dbm",
    "n_unique_epcs",
    "n_dropped_below_floor", "n_dropped_low_count", "n_dropped_ambiguous",
]

TAGS_HEADERS = ["tag_type", "label", "notes", "reference_image"]

CONFIG_HEADERS = ["key", "value"]

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
FAIL_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


def _ensure_sheet(wb: Workbook, name: str, headers: list[str]):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for col, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    return ws


def open_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, headers in [
        ("Summary", SUMMARY_HEADERS),
        ("Trials", TRIAL_HEADERS),
        ("Raw", RAW_HEADERS),
        ("Tags", TAGS_HEADERS),
        ("Config", CONFIG_HEADERS),
    ]:
        _ensure_sheet(wb, name, headers)
    return wb


def append_trial_to_workbook(wb: Workbook, trial: TrialResult) -> None:
    ws_trial = _ensure_sheet(wb, "Trials", TRIAL_HEADERS)
    ws_trial.append([
        trial.session_id, trial.trial_uid, trial.started_at_iso,
        trial.tag_type, trial.layout, trial.power_mw, trial.condition,
        trial.expected_ant0, trial.expected_ant1,
        trial.total_windows,
        round(trial.ant0_attribution_pct, 2),
        round(trial.ant1_attribution_pct, 2),
        trial.ant0_inferred_epc, trial.ant1_inferred_epc,
        trial.cross_read_count, trial.miss_read_count,
        round(trial.mean_rate_0_hz, 2),
        round(trial.mean_rate_1_hz, 2),
        round(trial.mean_rssi_ant0_dbm, 2) if trial.mean_rssi_ant0_dbm == trial.mean_rssi_ant0_dbm else "",
        round(trial.mean_rssi_ant1_dbm, 2) if trial.mean_rssi_ant1_dbm == trial.mean_rssi_ant1_dbm else "",
        trial.total_dropped_below_floor,
        trial.total_dropped_low_count,
        trial.total_dropped_ambiguous,
        trial.csv_path,
        trial.notes,
    ])

    ws_raw = _ensure_sheet(wb, "Raw", RAW_HEADERS)
    for w in trial.windows:
        ws_raw.append([
            trial.trial_uid, w.window_idx,
            w.t_start_unix, w.t_end_unix, w.power_mw, w.window_ms,
            w.scans_0, w.scans_1, w.rate_0_hz, w.rate_1_hz,
            ";".join(w.attr_ant0_epcs),
            ";".join(f"{x:.1f}" for x in w.attr_ant0_rssis),
            ";".join(w.attr_ant1_epcs),
            ";".join(f"{x:.1f}" for x in w.attr_ant1_rssis),
            w.n_unique_epcs,
            w.n_dropped_below_floor, w.n_dropped_low_count, w.n_dropped_ambiguous,
        ])


def rebuild_summary_sheet(wb: Workbook, cfg: Config,
                          setup_photo_lookup: dict[tuple[str, str, int, str], Path]) -> None:
    """Recompute Summary from every row currently on the Trials sheet.

    This is idempotent and re-runs every time we save, so cells reflect
    everything that has been recorded so far across all sessions.
    """
    ws_trial = wb["Trials"]
    rows = list(ws_trial.iter_rows(min_row=2, values_only=True))
    cells: dict[tuple[str, str, int, str], list[dict]] = defaultdict(list)
    for r in rows:
        if not r or r[0] is None:
            continue
        rec = dict(zip(TRIAL_HEADERS, r))
        key = (
            rec["tag_type"], rec["layout"], int(rec["power_mw"]), rec["condition"]
        )
        cells[key].append(rec)

    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    ws = wb.create_sheet("Summary", 0)
    for col, h in enumerate(SUMMARY_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for col in range(1, len(SUMMARY_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    row_idx = 2
    for tag in cfg.tag_types:
        for layout in cfg.layouts:
            for power in cfg.powers_mw:
                for cond in cfg.conditions:
                    key = (tag.id, layout.id, power, cond.id)
                    trials = cells.get(key, [])
                    if not trials:
                        continue
                    n_trials = len(trials)
                    n_windows = sum(int(t["total_windows"]) for t in trials)
                    ant0_pct = statistics.fmean(float(t["ant0_attribution_%"]) for t in trials)
                    ant1_pct = statistics.fmean(float(t["ant1_attribution_%"]) for t in trials)
                    cross = sum(int(t["cross_read_count"]) for t in trials)
                    miss = sum(int(t["miss_read_count"]) for t in trials)
                    expected_slots_per_trial = (
                        (1 if layout.expected["ant0"] else 0)
                        + (1 if layout.expected["ant1"] else 0)
                    )
                    total_expected_slots = expected_slots_per_trial * n_windows
                    miss_pct = (100.0 * miss / total_expected_slots) if total_expected_slots else 0.0
                    cross_pct = (100.0 * cross / max(n_windows, 1)) if n_windows else 0.0

                    rate0 = statistics.fmean(float(t["mean_rate_0_hz"]) for t in trials)
                    rate1 = statistics.fmean(float(t["mean_rate_1_hz"]) for t in trials)
                    rssi0_vals = [float(t["mean_rssi_ant0_dbm"]) for t in trials if t["mean_rssi_ant0_dbm"] not in ("", None)]
                    rssi1_vals = [float(t["mean_rssi_ant1_dbm"]) for t in trials if t["mean_rssi_ant1_dbm"] not in ("", None)]
                    rssi0 = statistics.fmean(rssi0_vals) if rssi0_vals else None
                    rssi1 = statistics.fmean(rssi1_vals) if rssi1_vals else None

                    ws.cell(row=row_idx, column=1, value=tag.id)
                    ws.cell(row=row_idx, column=2, value=layout.id)
                    ws.cell(row=row_idx, column=3, value=power)
                    ws.cell(row=row_idx, column=4, value=cond.id)
                    ws.cell(row=row_idx, column=5, value=n_trials)
                    ws.cell(row=row_idx, column=6, value=n_windows)
                    c_ant0 = ws.cell(row=row_idx, column=7, value=round(ant0_pct, 2))
                    c_ant1 = ws.cell(row=row_idx, column=8, value=round(ant1_pct, 2))
                    c_cross = ws.cell(row=row_idx, column=9, value=round(cross_pct, 2))
                    c_miss = ws.cell(row=row_idx, column=10, value=round(miss_pct, 2))
                    ws.cell(row=row_idx, column=11, value=round(rate0, 2))
                    ws.cell(row=row_idx, column=12, value=round(rate1, 2))
                    ws.cell(row=row_idx, column=13, value=round(rssi0, 2) if rssi0 is not None else "")
                    ws.cell(row=row_idx, column=14, value=round(rssi1, 2) if rssi1 is not None else "")

                    if layout.expected["ant0"] and ant0_pct < cfg.thresholds.attribution_rate_min_pct:
                        c_ant0.fill = FAIL_FILL
                    if layout.expected["ant1"] and ant1_pct < cfg.thresholds.attribution_rate_min_pct:
                        c_ant1.fill = FAIL_FILL
                    if cross_pct > cfg.thresholds.cross_read_rate_max_pct:
                        c_cross.fill = FAIL_FILL
                    if miss_pct > 0.0:
                        c_miss.fill = FAIL_FILL

                    photo = setup_photo_lookup.get(key)
                    if photo and photo.exists():
                        ws.row_dimensions[row_idx].height = 80
                        try:
                            embed_thumbnail(ws, photo, row_idx, col=15, max_px=120)
                        except Exception as e:
                            ws.cell(row=row_idx, column=15, value=f"<photo error: {e}>")
                    elif photo:
                        ws.cell(row=row_idx, column=15, value=str(photo))
                    row_idx += 1


def embed_thumbnail(ws, path: Path, row: int, col: int, max_px: int = 120) -> None:
    with PILImage.open(path) as im:
        im.thumbnail((max_px, max_px))
        buf = io.BytesIO()
        im.convert("RGB").save(buf, format="PNG")
        buf.seek(0)
    img = XLImage(buf)
    img.anchor = f"{get_column_letter(col)}{row}"
    ws.add_image(img)


def rebuild_tags_sheet(wb: Workbook, cfg: Config) -> None:
    if "Tags" in wb.sheetnames:
        wb.remove(wb["Tags"])
    ws = wb.create_sheet("Tags")
    for col, h in enumerate(TAGS_HEADERS + ["image"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = 22
    for i, t in enumerate(cfg.tag_types, start=2):
        ws.cell(row=i, column=1, value=t.id)
        ws.cell(row=i, column=2, value=t.label)
        ws.cell(row=i, column=3, value=t.notes)
        ws.cell(row=i, column=4, value=t.reference_image)
        if t.reference_image:
            img_path = (BENCH_DIR / t.reference_image).resolve()
            if img_path.exists():
                ws.row_dimensions[i].height = 100
                try:
                    embed_thumbnail(ws, img_path, i, col=5, max_px=140)
                except Exception as e:
                    ws.cell(row=i, column=5, value=f"<image error: {e}>")


def write_config_sheet(wb: Workbook, cfg: Config, session_id: str,
                       extra: dict[str, Any]) -> None:
    if "Config" in wb.sheetnames:
        wb.remove(wb["Config"])
    ws = wb.create_sheet("Config")
    ws.cell(row=1, column=1, value="key").font = HEADER_FONT
    ws.cell(row=1, column=2, value="value").font = HEADER_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    rows = [
        ("session_id", session_id),
        ("ssh_host", cfg.ssh.host),
        ("ssh_user", cfg.ssh.user),
        ("remote_workdir", cfg.ssh.remote_workdir),
        ("remote_binary", cfg.ssh.remote_binary),
        ("trial_duration_s", cfg.measurement.trial_duration_s),
        ("decision_window_ms", cfg.measurement.decision_window_ms),
        ("powers_mw", ", ".join(str(p) for p in cfg.powers_mw)),
        ("conditions", ", ".join(c.id for c in cfg.conditions)),
        ("layouts", ", ".join(lay.id for lay in cfg.layouts)),
        ("tag_types", ", ".join(t.id for t in cfg.tag_types)),
        ("attribution_threshold_pct", cfg.thresholds.attribution_rate_min_pct),
        ("cross_read_threshold_pct", cfg.thresholds.cross_read_rate_max_pct),
    ]
    for k, v in extra.items():
        rows.append((k, v))
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)


# ----------------------------------------------------------------------
# Interactive harness.
# ----------------------------------------------------------------------
def prompt_yes_no(question: str, default: bool = True) -> bool:
    suffix = " [Y/n] " if default else " [y/N] "
    while True:
        ans = input(question + suffix).strip().lower()
        if not ans:
            return default
        if ans in ("y", "yes"):
            return True
        if ans in ("n", "no"):
            return False
        print("  please answer y or n")


def prompt_choice(question: str, options: list[tuple[str, str]]) -> str:
    """Returns the id of the chosen option."""
    print(question)
    for i, (oid, label) in enumerate(options, 1):
        print(f"  {i}) [{oid}] {label}")
    while True:
        ans = input("> ").strip().lower()
        for oid, _ in options:
            if ans == oid:
                return oid
        if ans.isdigit():
            i = int(ans)
            if 1 <= i <= len(options):
                return options[i - 1][0]
        print("  invalid choice — type the id or the number")


def print_banner(title: str) -> None:
    bar = "=" * (len(title) + 4)
    print(f"\n{bar}\n  {title}\n{bar}")


def print_scenario(tag: TagType, layout: Layout, power: int, cond: Condition) -> None:
    print()
    print("-" * 72)
    print(f"  Tag type : {tag.label}")
    print(f"  Layout   : {layout.label}")
    print(f"  Power    : {power} mW")
    print(f"  Condition: {cond.label}")
    print("-" * 72)
    print(f"  Setup    : {layout.operator_prompt}")
    print(f"  Liquid   : {cond.operator_prompt}")
    print("-" * 72)


def print_trial_summary(t: TrialResult) -> None:
    print()
    print(f"  Trial UID            : {t.trial_uid}")
    print(f"  Decision windows     : {t.total_windows}")
    print(f"  Ant0 attribution     : {t.ant0_attribution_pct:6.2f}%   inferred EPC: {t.ant0_inferred_epc or '-'}")
    print(f"  Ant1 attribution     : {t.ant1_attribution_pct:6.2f}%   inferred EPC: {t.ant1_inferred_epc or '-'}")
    print(f"  Miss-reads           : {t.miss_read_count}")
    print(f"  Cross-reads          : {t.cross_read_count}")
    print(f"  Scan rate ant0/ant1  : {t.mean_rate_0_hz:.1f} Hz / {t.mean_rate_1_hz:.1f} Hz")
    rssi0 = f"{t.mean_rssi_ant0_dbm:.1f}" if t.mean_rssi_ant0_dbm == t.mean_rssi_ant0_dbm else "-"
    rssi1 = f"{t.mean_rssi_ant1_dbm:.1f}" if t.mean_rssi_ant1_dbm == t.mean_rssi_ant1_dbm else "-"
    print(f"  Mean RSSI ant0/ant1  : {rssi0} dBm / {rssi1} dBm")
    print(f"  Dropped below floor  : {t.total_dropped_below_floor}")
    print(f"  Dropped low count    : {t.total_dropped_low_count}")
    print(f"  Dropped ambiguous    : {t.total_dropped_ambiguous}")
    print(f"  CSV saved to         : {t.csv_path}")


def make_trial_uid(tag_id: str, layout_id: str, power: int, cond_id: str) -> str:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{tag_id}_{layout_id}_{power}mW_{cond_id}_{ts}"


def ask_setup_photo(scenario_key: tuple[str, str, int, str]) -> Path | None:
    """Let the operator drop a phone photo into bench/photos/setups/<key>.* ."""
    tag_id, layout_id, power, cond_id = scenario_key
    base = BENCH_DIR / "photos" / "setups" / f"{tag_id}_{layout_id}_{power}mW_{cond_id}"
    print(f"\n  If you want to attach a setup photo, drop it as:")
    print(f"    {base}.jpg  (or .png / .jpeg)")
    if not prompt_yes_no("  Photo ready?", default=False):
        return None
    for ext in (".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG"):
        p = base.with_suffix(ext)
        if p.exists():
            print(f"  using {p.name}")
            return p
    print("  no photo file found — skipping")
    return None


def collect_setup_photos(cfg: Config) -> dict[tuple[str, str, int, str], Path]:
    """Pre-scan photos/setups/ so already-dropped photos are picked up on re-runs."""
    out: dict[tuple[str, str, int, str], Path] = {}
    setup_dir = BENCH_DIR / "photos" / "setups"
    if not setup_dir.exists():
        return out
    for tag in cfg.tag_types:
        for layout in cfg.layouts:
            for power in cfg.powers_mw:
                for cond in cfg.conditions:
                    base = setup_dir / f"{tag.id}_{layout.id}_{power}mW_{cond.id}"
                    for ext in (".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG"):
                        p = base.with_suffix(ext)
                        if p.exists():
                            out[(tag.id, layout.id, power, cond.id)] = p
                            break
    return out


# ----------------------------------------------------------------------
# Main entry point.
# ----------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="Beer-pour RFID test orchestrator")
    ap.add_argument("--config", default=str(BENCH_DIR / "config.yaml"), help="path to config.yaml")
    ap.add_argument("--tag-type", default=None, help="run only this tag_type (default: ask)")
    args = ap.parse_args()

    cfg_path = Path(args.config).resolve()
    cfg = load_config(cfg_path)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    session_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    session_dir = RESULTS_DIR / f"session_{session_id}"
    session_dir.mkdir(parents=True, exist_ok=True)

    print_banner("Beer-pour RFID reliability harness")
    print(f"Config        : {cfg_path}")
    print(f"Session dir   : {session_dir}")
    print(f"Report (XLSX) : {REPORT_XLSX}")
    print(f"SSH target    : {cfg.ssh.user}@{cfg.ssh.host}:{cfg.ssh.port}")

    if args.tag_type:
        tag_choice = args.tag_type
    elif len(cfg.tag_types) == 1:
        tag_choice = cfg.tag_types[0].id
    else:
        tag_choice = prompt_choice(
            "Which tag type are you testing this session?",
            [(t.id, t.label) for t in cfg.tag_types],
        )
    chosen_tag = next((t for t in cfg.tag_types if t.id == tag_choice), None)
    if chosen_tag is None:
        print(f"ERROR: unknown tag_type '{tag_choice}'")
        return 2
    print(f"Tag type      : {chosen_tag.label}")

    runner = RemoteRunner(cfg.ssh)
    try:
        runner.connect()
    except Exception as e:
        print(f"\nERROR: cannot SSH into {cfg.ssh.host}: {e}")
        print("Check config.yaml `ssh:` block (host/user/key_path/password).")
        return 3

    try:
        remote_workdir_abs = runner.expand_remote_workdir()
        print(f"Remote workdir: {remote_workdir_abs}")
    except Exception as e:
        print(f"ERROR resolving remote_workdir: {e}")
        runner.close()
        return 4

    wb = open_or_create_workbook(REPORT_XLSX)
    setup_photos = collect_setup_photos(cfg)

    appended_count = 0
    aborted = False

    try:
        for layout in cfg.layouts:
            for power in cfg.powers_mw:
                for cond in cfg.conditions:
                    key = (chosen_tag.id, layout.id, power, cond.id)

                    while True:
                        print_scenario(chosen_tag, layout, power, cond)
                        if not prompt_yes_no("  Run a trial for this cell?", default=True):
                            print("  skipping this cell.")
                            break

                        input("\n  >>> Press ENTER when the setup is in place (and pour, if wet) <<<")

                        trial_uid = make_trial_uid(chosen_tag.id, layout.id, power, cond.id)
                        remote_csv = posixpath.join(
                            remote_workdir_abs, f".trial_{trial_uid}.csv"
                        )
                        local_csv = session_dir / f"{trial_uid}.csv"

                        print(f"\n  [run] {chosen_tag.id} | {layout.id} | {power} mW | {cond.id}")
                        t_started = time.time()
                        try:
                            status, out, err = runner.run_trial(
                                power_mw=power,
                                duration_s=cfg.measurement.trial_duration_s,
                                window_ms=cfg.measurement.decision_window_ms,
                                remote_csv_path=remote_csv,
                            )
                        except Exception as e:
                            print(f"  ERROR: remote run failed: {e}")
                            if not prompt_yes_no("  Retry this trial?", default=True):
                                break
                            continue
                        t_elapsed = time.time() - t_started
                        print(f"  [run] exit={status} elapsed={t_elapsed:.1f}s")
                        if err.strip():
                            print(f"  [stderr] {err.strip()}")
                        if status != 0:
                            print("  binary returned non-zero exit. Retry?")
                            if not prompt_yes_no("  Retry?", default=True):
                                break
                            continue

                        try:
                            runner.fetch(remote_csv, local_csv)
                        except Exception as e:
                            print(f"  ERROR fetching CSV: {e}")
                            if not prompt_yes_no("  Retry this trial?", default=True):
                                break
                            continue

                        windows = parse_csv(local_csv)
                        if not windows:
                            print("  WARNING: CSV is empty (reader produced no windows).")

                        trial = TrialResult(
                            session_id=session_id,
                            trial_uid=trial_uid,
                            started_at_iso=datetime.fromtimestamp(t_started, tz=timezone.utc).isoformat(),
                            tag_type=chosen_tag.id,
                            layout=layout.id,
                            power_mw=power,
                            condition=cond.id,
                            expected_ant0=bool(layout.expected.get("ant0", False)),
                            expected_ant1=bool(layout.expected.get("ant1", False)),
                            csv_path=str(local_csv.relative_to(BENCH_DIR)),
                            windows=windows,
                        )
                        compute_metrics(trial)
                        print_trial_summary(trial)

                        if prompt_yes_no("\n  Append this trial to the report?", default=True):
                            append_trial_to_workbook(wb, trial)
                            appended_count += 1
                            if key not in setup_photos:
                                photo = ask_setup_photo(key)
                                if photo:
                                    setup_photos[key] = photo
                            wb.save(REPORT_XLSX)
                            print(f"  appended. report has {appended_count} new trial(s) so far in this session.")
                        else:
                            print("  discarded.")

                        if not prompt_yes_no("\n  Run another trial in THIS cell?", default=False):
                            break

                        if cfg.measurement.inter_trial_pause_s > 0:
                            time.sleep(cfg.measurement.inter_trial_pause_s)

    except KeyboardInterrupt:
        print("\n\n  Ctrl+C caught — finishing up cleanly.")
        aborted = True
    finally:
        runner.close()
        try:
            rebuild_tags_sheet(wb, cfg)
            rebuild_summary_sheet(wb, cfg, setup_photos)
            write_config_sheet(wb, cfg, session_id, extra={
                "session_dir": str(session_dir),
                "appended_trials_this_session": appended_count,
                "aborted": aborted,
            })
            wb.save(REPORT_XLSX)
        except Exception as e:
            print(f"  WARNING: could not finalise workbook: {e}")

    print_banner("Done")
    print(f"  Trials appended this session: {appended_count}")
    print(f"  Report: {REPORT_XLSX}")
    print(f"  CSVs:   {session_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
